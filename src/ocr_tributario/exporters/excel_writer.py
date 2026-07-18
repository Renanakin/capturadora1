"""Exportador Excel (Fase 9)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ocr_tributario.config.schema import Settings
from ocr_tributario.models.invoice import InvoiceRecord

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
REVISION_FILL = PatternFill("solid", fgColor="FFC7CE")  # rojo claro para QUARANTINE/REJECTED

# Tipos nativos por columna (formato preestablecido en settings.yaml)
_NATIVE_TYPE: dict[str, str] = {
    "Archivo": "string",
    "Mes": "string",
    "Fecha": "date",
    "Nro Boleta Factura": "int",
    "PROVEEDOR": "string",
    "RUT": "string",
    "Total": "int",
    "Descripción del gasto": "string",
    "Observaciones": "string",
}


def _cast_for_column(col: str, value):
    """Devuelve un valor con tipo nativo (int/date) cuando es posible,
    para que Excel aplique formato numérico/fecha correcto al abrirlo."""
    if value is None or value == "":
        return None
    target = _NATIVE_TYPE.get(col, "string")
    try:
        if target == "int":
            return int(value)
        if target == "date":
            if isinstance(value, str) and len(value) >= 10:
                return datetime.fromisoformat(value[:10]).date()
    except (ValueError, TypeError):
        return value
    return value


def _row_for(record: InvoiceRecord, columns: list[str]) -> list:
    mapping = {
        "Archivo": record.archivo_origen,
        "Mes": record.mes or "",
        "Fecha": record.fecha or "",
        "Nro Boleta Factura": record.nro_documento if record.nro_documento is not None else "",
        "PROVEEDOR": record.proveedor or "",
        "RUT": record.rut or "",
        "Total": record.total if record.total is not None else "",
        "Descripción del gasto": record.descripcion or "",
        "Observaciones": record.observaciones or "",
    }
    return [_cast_for_column(col, mapping.get(col, "")) for col in columns]


def _apply_column_widths(ws, columns: list[str]) -> None:
    for i, col in enumerate(columns, start=1):
        col_letter = get_column_letter(i)
        # Ancho según contenido esperado
        width = max(14, len(col) + 4)
        if col == "PROVEEDOR":
            width = 38
        elif col == "RUT":
            width = 16
        elif col == "Fecha":
            width = 14
        elif col == "Mes":
            width = 10
        elif col == "Nro Boleta Factura":
            width = 18
        elif col == "Total":
            width = 14
        elif col == "Descripción del gasto":
            width = 32
        elif col == "Observaciones":
            width = 32
        ws.column_dimensions[col_letter].width = width


def _apply_cell_format(ws, columns: list[str]) -> None:
    """Aplica formato numérico a columnas int y formato fecha a columna Fecha."""
    for i, col in enumerate(columns, start=1):
        col_letter = get_column_letter(i)
        if col == "Total":
            for cell in ws[col_letter][1:]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
        elif col == "Nro Boleta Factura":
            for cell in ws[col_letter][1:]:
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal="right")
        elif col == "Fecha":
            for cell in ws[col_letter][1:]:
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal="center")


def _format_header(ws, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_records(
    records: list[InvoiceRecord],
    output_path: Path | None,
    settings: Settings,
) -> Path:
    """Escribe el Excel de rendición con dos hojas: Procesados + Revisión Manual."""
    output_dir: Path = settings.paths.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        period = datetime.now().strftime("%Y-%m")
        output_path = output_dir / f"Rendicion_Gastos_OCR_{period}.xlsx"

    wb = Workbook()
    ws_ok = wb.active
    ws_ok.title = settings.excel.sheet_procesados

    columns = settings.excel.template_columns
    ws_ok.append(columns)
    _format_header(ws_ok)
    if settings.excel.freeze_header:
        ws_ok.freeze_panes = "A2"

    ok_records = [r for r in records if r.estado == "OK"]
    for rec in ok_records:
        ws_ok.append(_row_for(rec, columns))

    _apply_column_widths(ws_ok, columns)
    _apply_cell_format(ws_ok, columns)

    # Hoja 2: revisión manual — mismas 9 columnas + estado + motivo al inicio
    ws_rev = wb.create_sheet(settings.excel.sheet_revision)
    rev_columns = ["estado", "motivo_revision", *columns]
    ws_rev.append(rev_columns)
    _format_header(ws_rev)
    if settings.excel.freeze_header:
        ws_rev.freeze_panes = "A2"

    for rec in records:
        if rec.estado in ("QUARANTINE", "REJECTED"):
            row = [
                rec.estado,
                rec.motivo_revision or "",
                *_row_for(rec, columns),
            ]
            ws_rev.append(row)
            # Resaltar fila en rojo claro
            for cell in ws_rev[ws_rev.max_row]:
                cell.fill = REVISION_FILL

    rev_widths = [14, 60, *[22 if c == "PROVEEDOR" else 14 for c in columns]]
    for i, w in enumerate(rev_widths, start=1):
        ws_rev.column_dimensions[get_column_letter(i)].width = w
    _apply_cell_format(ws_rev, columns)

    wb.save(output_path)
    logger.info(f"Excel OK ({len(ok_records)} registros) -> {output_path}")
    return output_path