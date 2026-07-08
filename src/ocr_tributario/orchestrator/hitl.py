"""Cola de revisión HITL (Fase 8, simplificado)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ocr_tributario.config.schema import Settings
from ocr_tributario.models.invoice import InvoiceRecord


def write_quarantine_excel(records: list[InvoiceRecord], settings: Settings) -> Path:
    """Genera un Excel de revisión con todos los documentos en cuarentena."""
    out_dir: Path = settings.paths.quarantine_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"revisar_manual_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Revisión"

    columns = [
        "archivo_origen", "estado", "motivo_revision",
        "fecha", "nro_documento", "proveedor", "rut", "total", "observaciones",
    ]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFC000")
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        ws.append([
            rec.archivo_origen,
            rec.estado,
            rec.motivo_revision or "",
            rec.fecha or "",
            rec.nro_documento or "",
            rec.proveedor or "",
            rec.rut or "",
            rec.total if rec.total is not None else "",
            rec.observaciones or "",
        ])

    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(col) + 4)

    wb.save(out_path)
    logger.info(f"HITL: {len(records)} docs en {out_path}")
    return out_path